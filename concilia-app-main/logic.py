import pandas as pd
from rapidfuzz import fuzz
from typing import Optional
import datetime
from io import BytesIO

def _limpiar_monto(s: pd.Series) -> pd.Series:
    """Función auxiliar para limpiar los montos y asegurar que sean numéricos."""
    s_str = s.astype(str)
    s_str = s_str.str.replace(r'[^\d,\.\-]', '', regex=True)
    
    def format_number(val):
        if pd.isna(val) or val == '': 
            return val
        val = str(val)
        
        last_comma = val.rfind(',')
        last_dot = val.rfind('.')
        
        if last_comma > -1 and last_dot > -1:
            if last_comma > last_dot:
                # Formato 1.234,56
                val = val.replace('.', '').replace(',', '.')
            else:
                # Formato 1,234.56
                val = val.replace(',', '')
        elif last_comma > -1:
            if val.count(',') > 1 or (len(val) - last_comma == 4):
                val = val.replace(',', '')
            else:
                val = val.replace(',', '.')
        elif last_dot > -1:
            if val.count('.') > 1 or (len(val) - last_dot == 4):
                val = val.replace('.', '')
                
        return val

    s_str = s_str.apply(format_number)
    num = pd.to_numeric(s_str, errors='coerce')
    return num, num.abs().round(2)

def procesar_conciliacion(
    df_banco: pd.DataFrame, 
    df_contable: pd.DataFrame, 
    col_fecha_banco: Optional[str], 
    col_monto_banco: str, 
    col_desc_banco: Optional[str], 
    col_fecha_conta: Optional[str], 
    col_monto_conta: str, 
    col_desc_conta: Optional[str],
    tolerancia_dias: int = 3
):
    # Crear copias para no alterar los originales y evitar SettingWithCopyWarning
    banco = df_banco.copy()
    contable = df_contable.copy()
    
    # Limpieza de montos y forzado a numérico utilizando .loc
    banco.loc[:, '_monto_raw_num'], banco.loc[:, '_monto_clean'] = _limpiar_monto(banco[col_monto_banco])
    contable.loc[:, '_monto_raw_num'], contable.loc[:, '_monto_clean'] = _limpiar_monto(contable[col_monto_conta])
    
    # Generar identificadores únicos
    banco.loc[:, '_id'] = range(len(banco))
    contable.loc[:, '_id'] = range(len(contable))
    
    banco['Estado_Color'] = 'Rojo'
    contable['Estado_Color'] = 'Rojo'
    banco['Similitud_%'] = None
    contable['Similitud_%'] = None
    
    matched_b = set()
    matched_l = set()
    coincidencias_altas = []
    coincidencias_bajas = []
    
    # Agrupamos por monto para optimizar la búsqueda
    for monto, group_b in banco.groupby('_monto_clean'):
        if pd.isna(monto): 
            continue
        
        group_l = contable[contable['_monto_clean'] == monto]
        if group_l.empty: 
            continue
        
        for _, row_b in group_b.iterrows():
            if row_b['_id'] in matched_b: 
                continue
            
            best_match = None
            best_score = -1
            
            # Extraemos la fecha del banco si existe
            fb = None
            if col_fecha_banco and not pd.isna(row_b.get(col_fecha_banco)):
                try:
                    val = row_b[col_fecha_banco]
                    fb = pd.to_datetime(str(val).strip(), dayfirst=True) if isinstance(val, str) else pd.to_datetime(val)
                except:
                    pass
            
            for _, row_l in group_l.iterrows():
                if row_l['_id'] in matched_l: 
                    continue
                
                # Validación de Signos: No permitir cruces entre positivos y negativos
                raw_b = row_b['_monto_raw_num']
                raw_l = row_l['_monto_raw_num']
                if (raw_b > 0 and raw_l < 0) or (raw_b < 0 and raw_l > 0):
                    continue
                
                # Validación de Fechas (Ventana de ±3 días)
                if fb and col_fecha_conta and not pd.isna(row_l.get(col_fecha_conta)):
                    try:
                        val = row_l[col_fecha_conta]
                        fl = pd.to_datetime(str(val).strip(), dayfirst=True) if isinstance(val, str) else pd.to_datetime(val)
                        if abs((fb - fl).days) > tolerancia_dias:
                            continue  # Fuera de la ventana de fechas
                    except:
                        pass
                
                # Evaluación Fuzzy si hay descripciones disponibles y no son 'Ninguno'
                if col_desc_banco and col_desc_conta:
                    desc_b = str(row_b[col_desc_banco])
                    desc_l = str(row_l[col_desc_conta])
                    score = fuzz.token_sort_ratio(desc_b, desc_l)
                else:
                    score = 100
                
                if score > best_score:
                    best_score = score
                    best_match = row_l
            
            # Si se encuentra una coincidencia de monto exacta, se asume match.
            # El fuzzy score (similitud) sirve para decidir en caso de empate de montos.
            if best_match is not None:
                matched_b.add(row_b['_id'])
                matched_l.add(best_match['_id'])
                
                color = 'Verde' if best_score > 35 else 'Azul'
                banco.loc[banco['_id'] == row_b['_id'], 'Estado_Color'] = color
                contable.loc[contable['_id'] == best_match['_id'], 'Estado_Color'] = color
                banco.loc[banco['_id'] == row_b['_id'], 'Similitud_%'] = best_score
                contable.loc[contable['_id'] == best_match['_id'], 'Similitud_%'] = best_score
                
                res = {}
                for c in df_banco.columns:
                    if c not in ['_monto_clean', '_id', '_monto_raw_num']:
                        res[f'{c}_banco'] = row_b[c]
                for c in df_contable.columns:
                    if c not in ['_monto_clean', '_id', '_monto_raw_num']:
                        res[f'{c}_contable'] = best_match[c]
                res['Similitud_Detalle_%'] = best_score
                
                if best_score > 35:
                    coincidencias_altas.append(res)
                else:
                    coincidencias_bajas.append(res)
                
    df_altas = pd.DataFrame(coincidencias_altas)
    df_bajas = pd.DataFrame(coincidencias_bajas)
    
    cols_to_drop = ['_id', '_monto_clean', '_monto_raw_num', 'Estado_Color', 'Similitud_%']
    solo_banco = banco[~banco['_id'].isin(matched_b)].drop(columns=cols_to_drop, errors='ignore')
    solo_contable = contable[~contable['_id'].isin(matched_l)].drop(columns=cols_to_drop, errors='ignore')
    
    banco_original = banco.drop(columns=['_id', '_monto_clean', '_monto_raw_num'], errors='ignore')
    contable_original = contable.drop(columns=['_id', '_monto_clean', '_monto_raw_num'], errors='ignore')
    
    return df_altas, df_bajas, solo_banco, solo_contable, banco_original, contable_original

def generar_excel_coloreado(df: pd.DataFrame) -> BytesIO:
    from openpyxl.styles import PatternFill
    output = BytesIO()
    
    color_map = {
        'Verde': 'C6EFCE',
        'Azul': 'B8CCE4',
        'Rojo': 'FFC7CE'
    }
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name='Resultados')
        workbook = writer.book
        worksheet = writer.sheets['Resultados']
        
        estado_col_idx = None
        for idx, col in enumerate(df.columns, start=1):
            if col == 'Estado_Color':
                estado_col_idx = idx
                break
                
        if estado_col_idx:
            # Color rows. with header=False, data starts at row 1
            for row_idx in range(1, len(df) + 1):
                estado = worksheet.cell(row=row_idx, column=estado_col_idx).value
                if estado in color_map:
                    fill_color = color_map[estado]
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = fill
            
            # Remove Estado_Color column
            worksheet.delete_cols(estado_col_idx)
            
    output.seek(0)
    return output
