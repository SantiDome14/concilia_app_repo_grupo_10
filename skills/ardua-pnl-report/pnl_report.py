"""
Reporte de P&L — módulo principal del Skill Ardua.

Transforma MASTER.xlsx en un reporte HTML print-ready de P&L siguiendo el layout
canónico del Report 9 (14/04/2026).

Uso programático:
    from pnl_report import generate_report
    path = generate_report('/path/to/MASTER.xlsx', 'daily', output_dir='./out')

Uso CLI: ver generate.py
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# ── Constantes ────────────────────────────────────────────────────────────────

SKIP_SHEETS = ('CRYPTO', 'Sheet1')
EXCLUDED_SEGMENTS = ('NO SEGMENT', 'INTERNAL TRANSFER', 'ARDUA CNT')
EXCLUDED_TYPE = 'CARD PURCHASE TXS'  # excluido desde el parseo en todas las agregaciones
REQUIRED_COLUMNS = ('Date', 'Partner', 'Item', 'Amount', 'Segment')

BANK_COLORS = {
    'BR USD': '#059669',
    'BR EUR': '#1D4ED8',
    'FV USD': '#B45309',
    'CV USD': '#7C3AED',
    'LRM USD': '#DB2777',
}
DEFAULT_BANK_COLOR = '#334155'


# ── Modelo ────────────────────────────────────────────────────────────────────

@dataclass
class Row:
    """Una fila normalizada del MASTER, post-parseo."""
    date: str  # YYYY-MM-DD
    partner: str
    item: str  # DEPOSIT / WITHDRAWAL / FX
    description: str
    segment: str
    client_name: str
    type: str
    amount: float
    ardua_fee: float
    ardua_fix: float
    onboarding: float
    notes: str
    currency: str
    tc_rate: float
    off_vol: bool
    off_rev: bool
    revenue: float
    volume: float
    sheet: str


# ── Errores ───────────────────────────────────────────────────────────────────

class PnLError(Exception):
    """Error de dominio del reporte de P&L. El mensaje es directo para el usuario."""
    pass


# ── Utilidades de parseo ──────────────────────────────────────────────────────

def _parse_date(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    # Try ISO
    try:
        return datetime.fromisoformat(s[:19]).date().isoformat()
    except ValueError:
        pass
    # Try DD-MM-YYYY or DD/MM/YYYY
    for sep in ('-', '/'):
        parts = s.split(sep)
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            d, m, y = parts
            if len(y) == 2:
                y = '20' + y
            try:
                return date(int(y), int(m), int(d)).isoformat()
            except ValueError:
                return None
    return None


def _parse_fee(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, str):
        s = v.strip()
        if '%' in s:
            try:
                return float(s.replace('%', '').replace(',', '.')) / 100
            except ValueError:
                return 0.0
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _parse_amount(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(',', '').replace(' ', ''))
        except ValueError:
            return 0.0
    return 0.0


def _should_skip_sheet(sheet_name: str) -> bool:
    su = sheet_name.upper()
    return any(s.upper() in su for s in SKIP_SHEETS)


def _extract_currency(sheet_name: str) -> str:
    """'BRIDGE USD' → 'USD', 'BR EUR' → 'EUR'. Fallback USD."""
    parts = sheet_name.strip().split()
    return parts[-1].upper() if len(parts) > 1 else 'USD'


# ── Parser ────────────────────────────────────────────────────────────────────

def parse_master(master_path: str | Path) -> list[Row]:
    """
    Parsea MASTER.xlsx y devuelve filas normalizadas.

    - Saltea hojas CRYPTO y Sheet1
    - Excluye filas con Type == 'CARD PURCHASE TXS' (desde el parseo)
    - Excluye filas sin fecha parseable
    - Normaliza strings con UPPER + TRIM (Segment, Partner, Item, Type)
    - Calcula revenue y volume por fila según reglas de pnl-session-context.md §4.1
    """
    path = Path(master_path)
    if not path.exists():
        raise PnLError(
            f"No encuentro MASTER.xlsx en `{path.parent}`. "
            "Verificá que el archivo esté en esa ubicación."
        )

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        raise PnLError(f"No pude abrir MASTER.xlsx: {e}")

    rows: list[Row] = []
    sheets_processed = 0

    for sheet_name in wb.sheetnames:
        if _should_skip_sheet(sheet_name):
            continue

        ws = wb[sheet_name]
        header: Optional[list] = None
        first_data = True

        for i, raw_row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = list(raw_row)
                continue
            # Skip fully empty rows
            if not any(c is not None for c in raw_row):
                continue

            d = dict(zip(header, raw_row))

            # Validate required columns (once per sheet, on first data row)
            if first_data:
                missing = [c for c in REQUIRED_COLUMNS if c not in header]
                if missing:
                    raise PnLError(
                        f"La hoja `{sheet_name}` no tiene la(s) columna(s) "
                        f"`{', '.join(missing)}`. Columnas obligatorias: "
                        f"{', '.join(REQUIRED_COLUMNS)}."
                    )
                first_data = False
                sheets_processed += 1

            date_str = _parse_date(d.get('Date'))
            if not date_str:
                continue  # row sin fecha parseable

            type_ = str(d.get('Type') or '').upper().strip()
            if type_ == EXCLUDED_TYPE:
                continue  # fix D-07: exclusión global desde el parseo

            partner = str(d.get('Partner') or '').upper().strip()
            item = str(d.get('Item') or '').upper().strip()
            segment = str(d.get('Segment') or 'NO SEGMENT').upper().strip()
            currency = _extract_currency(sheet_name)

            amount = _parse_amount(d.get('Amount'))
            ardua_fee = _parse_fee(d.get('Ardua Fee'))
            ardua_fix = _parse_fee(d.get('Ardua Fixed Cost'))
            onboarding = _parse_fee(d.get('Onboarding'))
            notes = str(d.get('Notes') or '').upper().strip()

            off_vol = 'OFF VOL' in notes
            off_rev = 'OFF REV' in notes

            # tc_rate: FX Rate column, or TC X.XXX in Notes, or 1
            tc = 1.0
            fx = d.get('FX Rate') or d.get('FX rate')
            if fx is not None:
                try:
                    tc = float(str(fx).replace(',', '.')) or 1.0
                except ValueError:
                    tc = 1.0
            else:
                m = re.search(r'TC\s*([\d,\.]+)', notes)
                if m:
                    try:
                        tc = float(m.group(1).replace(',', '.')) or 1.0
                    except ValueError:
                        tc = 1.0

            # Revenue and volume
            volume = 0.0 if off_vol else amount
            if off_rev:
                revenue = 0.0
            elif currency != 'USD' and tc != 1:
                revenue = amount * tc * ardua_fee + ardua_fix + onboarding
            else:
                revenue = amount * ardua_fee + ardua_fix + onboarding

            rows.append(Row(
                date=date_str,
                partner=partner,
                item=item,
                description=str(d.get('Description') or '').strip(),
                segment=segment,
                client_name=str(d.get('Client Name') or '').upper().strip(),
                type=type_,
                amount=round(amount, 2),
                ardua_fee=ardua_fee,
                ardua_fix=ardua_fix,
                onboarding=onboarding,
                notes=notes,
                currency=currency,
                tc_rate=tc,
                off_vol=off_vol,
                off_rev=off_rev,
                revenue=round(revenue, 2),
                volume=round(volume, 2),
                sheet=sheet_name,
            ))

    if sheets_processed == 0:
        raise PnLError(
            "MASTER.xlsx no contiene hojas con datos. "
            "Revisá que al menos una hoja de Partner tenga filas."
        )

    return rows


# ── Cómputo de períodos ───────────────────────────────────────────────────────

def resolve_period(rows: list[Row], report_type: str,
                   from_date: Optional[str] = None,
                   to_date: Optional[str] = None) -> tuple[str, str]:
    """Determina (from_date, to_date) según report_type y el dataset disponible."""
    if not rows:
        raise PnLError("No hay movimientos en MASTER.xlsx para generar el reporte.")

    last_date = max(r.date for r in rows)
    ld = date.fromisoformat(last_date)

    if report_type == 'daily':
        return last_date, last_date
    if report_type == 'weekly':
        return (ld - timedelta(days=6)).isoformat(), last_date
    if report_type == 'monthly':
        return (ld - timedelta(days=29)).isoformat(), last_date
    if report_type == 'custom':
        if not from_date or not to_date:
            raise PnLError(
                "El reporte custom requiere `from_date` y `to_date` en formato YYYY-MM-DD."
            )
        if from_date > to_date:
            raise PnLError(
                f"La fecha desde (`{from_date}`) es posterior a la fecha hasta "
                f"(`{to_date}`). Invertí el orden."
            )
        return from_date, to_date
    raise PnLError(
        f"Tipo de reporte inválido: `{report_type}`. "
        "Opciones: daily / weekly / monthly / custom."
    )


def _filter_range(rows: list[Row], from_d: str, to_d: str) -> list[Row]:
    return [r for r in rows if from_d <= r.date <= to_d]


def _prev_period(from_d: str, to_d: str) -> tuple[str, str]:
    """Mismo largo en días, inmediatamente anterior."""
    f = date.fromisoformat(from_d)
    t = date.fromisoformat(to_d)
    span_days = (t - f).days + 1
    pf = f - timedelta(days=span_days)
    pt = f - timedelta(days=1)
    return pf.isoformat(), pt.isoformat()


# ── Agregaciones ──────────────────────────────────────────────────────────────

def compute_aggregates(rows_all: list[Row], from_d: str, to_d: str) -> dict:
    """Calcula todas las métricas que necesita el renderer."""
    last_date = max(r.date for r in rows_all) if rows_all else to_d
    d = _filter_range(rows_all, from_d, to_d)

    # Previous period
    pf, pt = _prev_period(from_d, to_d)
    prev = _filter_range(rows_all, pf, pt)

    # 7D / 30D referenced to last_date
    ld = date.fromisoformat(last_date)
    d7_from = (ld - timedelta(days=6)).isoformat()
    d30_from = (ld - timedelta(days=29)).isoformat()
    d7 = _filter_range(rows_all, d7_from, last_date)
    d30 = _filter_range(rows_all, d30_from, last_date)

    # Totals
    rv = round(sum(r.revenue for r in d), 2)
    vl = round(sum(r.volume for r in d), 2)
    tx = len(d)
    prv = round(sum(r.revenue for r in prev), 2)
    pvl = round(sum(r.volume for r in prev), 2)

    # Days distinct
    dy = len({r.date for r in d}) or 1

    # Net flow
    dep = sum(r.volume for r in d if r.item == 'DEPOSIT')
    wdr = sum(r.volume for r in d if r.item == 'WITHDRAWAL')
    net_flow = round(dep - wdr, 2)

    # Largest txn
    largest = max((r.amount for r in d), default=0.0)

    # By partner (bank breakdown)
    by_partner: dict[str, dict] = {}
    for r in d:
        bp = by_partner.setdefault(r.partner, {'rv': 0.0, 'vl': 0.0, 'tx': 0, 'ccy': r.currency})
        bp['rv'] += r.revenue
        bp['vl'] += r.volume
        bp['tx'] += 1
    for bp in by_partner.values():
        bp['rv'] = round(bp['rv'], 2)
        bp['vl'] = round(bp['vl'], 2)

    # By type (rail)
    by_type: dict[str, dict] = {}
    for r in d:
        bt = by_type.setdefault(r.type or '—', {'rv': 0.0, 'tx': 0})
        bt['rv'] += r.revenue
        bt['tx'] += 1
    for bt in by_type.values():
        bt['rv'] = round(bt['rv'], 2)

    # By segment (excluding EXCLUDED_SEGMENTS from breakdown but counting in totals)
    by_seg: dict[str, dict] = {}
    for r in d:
        if r.segment in EXCLUDED_SEGMENTS:
            continue
        bs = by_seg.setdefault(r.segment, {'rv': 0.0, 'tx': 0})
        bs['rv'] += r.revenue
        bs['tx'] += 1
    for bs in by_seg.values():
        bs['rv'] = round(bs['rv'], 2)

    # By currency
    by_ccy: dict[str, dict] = {}
    for r in d:
        bc = by_ccy.setdefault(r.currency, {'rv': 0.0, 'vl': 0.0, 'tx': 0})
        bc['rv'] += r.revenue
        bc['vl'] += r.volume
        bc['tx'] += 1
    for bc in by_ccy.values():
        bc['rv'] = round(bc['rv'], 2)
        bc['vl'] = round(bc['vl'], 2)

    # All-time totals and averages (applied consistently — CARD PURCHASE TXS already excluded from rows_all)
    total_rv = sum(r.revenue for r in rows_all)
    total_vl = sum(r.volume for r in rows_all)
    total_tx = len(rows_all)

    def avg(vals_sum, n):
        return (vals_sum / n) if n else 0

    return {
        'from': from_d,
        'to': to_d,
        'last_date': last_date,
        'revenue': rv,
        'volume': vl,
        'transfers': tx,
        'prev_revenue': prv,
        'prev_volume': pvl,
        'prev_transfers': len(prev),
        'profit': rv,  # V1: cost = 0, profit = revenue
        'cost': 0.0,
        'days_distinct': dy,
        'net_flow': net_flow,
        'largest_txn': largest,
        'avg_rev_per_txn': round(rv / tx, 2) if tx else 0.0,
        'rev_vol_ratio': (rv / vl) if vl else 0.0,
        # 7D / 30D averages (per active day in that window)
        'avg_rev_7d': round(sum(r.revenue for r in d7) / (len({r.date for r in d7}) or 1), 2),
        'avg_vol_7d': round(sum(r.volume for r in d7) / (len({r.date for r in d7}) or 1), 2),
        'avg_tx_7d': round(len(d7) / (len({r.date for r in d7}) or 1), 1),
        'avg_rev_30d': round(sum(r.revenue for r in d30) / (len({r.date for r in d30}) or 1), 2),
        'avg_vol_30d': round(sum(r.volume for r in d30) / (len({r.date for r in d30}) or 1), 2),
        'avg_tx_30d': round(len(d30) / (len({r.date for r in d30}) or 1), 1),
        # All-time averages — consistent: rows_all already excludes CARD PURCHASE TXS
        'avg_rev_all': round(total_rv / dy, 2),
        'avg_vol_all': round(total_vl / dy, 2),
        'avg_tx_all': round(total_tx / dy, 1),
        'total_rev_all': round(total_rv, 2),
        # All-time averages per-row (for PPS benchmarks)
        'allrows_avg_rev': round(total_rv / total_tx, 2) if total_tx else 0.0,
        'allrows_rev_vol': (total_rv / total_vl) if total_vl else 0.0,
        # Breakdowns
        'by_partner': by_partner,
        'by_type': by_type,
        'by_seg': by_seg,
        'by_ccy': by_ccy,
    }


# ── Formato de números ────────────────────────────────────────────────────────

def _fD(n: float) -> str:
    """Dollar with 2 decimals and comma separators: $4,872.38"""
    return f"${n:,.2f}"


def _fM(n: float) -> str:
    """Magnitude: $2.64M, $130.85M, $-364.57K, else $4,872.38"""
    a = abs(n)
    if a >= 1e9:
        return f"${n/1e9:,.2f}B"
    if a >= 1e6:
        return f"${n/1e6:,.2f}M"
    if a >= 1e3:
        return f"${n/1e3:,.2f}K"
    return f"${n:,.2f}"


def _fP(n: float) -> str:
    """Signed percent: +0.20%, -44.40%. Uses regular minus for HTML."""
    sign = '+' if n > 0 else ('' if n < 0 else '')
    return f"{sign}{n:.2f}%"


def _f0(n: float) -> str:
    return f"{int(round(n)):,}"


def _delta(a: float, b: float) -> float:
    return (a - b) / abs(b) * 100 if b else 0.0


# ── Render ────────────────────────────────────────────────────────────────────

REPORT_LABELS = {
    'daily': 'DAILY P&L REPORT',
    'weekly': 'WEEKLY P&L REPORT',
    'monthly': 'MONTHLY P&L REPORT',
    'custom': 'CUSTOM P&L REPORT',
}


# ── Paginación automática ─────────────────────────────────────────────────────
# Alturas estimadas en mm para cada sección, basadas en el CSS actual.
# La hoja A4 tiene 297mm. Padding top 10mm + bottom 8mm = 18mm. Header ~10mm
# y footer ~7mm se repiten en cada hoja, dejando ~262mm útiles para secciones.

PAGE_USABLE_MM = 258            # espacio útil para secciones (post header/footer)
SECTION_H_KPIS = 22             # KPI row (3 cards, font 20px)
SECTION_H_PERF = 26             # Performance Summary (título + grid 4 cards)
SECTION_H_PERIOD = 30           # Period Comparison (título + tabla 3 filas)
SECTION_H_BR_BASE = 16          # Bank+Rails: altura base (título + card padding)
SECTION_H_BR_ROW = 4            # altura por fila en Bank Breakdown
SECTION_H_RAIL_ROW = 4          # altura por fila en Rails/Segments
SECTION_H_CCY_BASE = 16         # Currency: altura base (título + header de tabla)
SECTION_H_CCY_ROW = 5           # altura por fila de moneda


def _estimate_br_h(agg: dict) -> float:
    """Altura estimada del bloque Bank Breakdown + Rail & Society (dos columnas)."""
    n_banks = len(agg['by_partner'])
    n_rails = len(agg['by_type'])
    n_segs = min(5, len([s for s in agg['by_seg']]))
    # Columna Bank: n_banks filas
    bank_col = n_banks * SECTION_H_BR_ROW
    # Columna Rail & Society: n_rails + separador (~2mm) + n_segs + top_label (~3mm)
    rail_col = n_rails * SECTION_H_RAIL_ROW + 2 + n_segs * SECTION_H_RAIL_ROW + 3
    return SECTION_H_BR_BASE + max(bank_col, rail_col)


def _estimate_ccy_h(agg: dict) -> float:
    """Altura estimada del Currency Breakdown. 0 si no se muestra."""
    if len(agg['by_ccy']) <= 1:
        return 0.0
    return SECTION_H_CCY_BASE + len(agg['by_ccy']) * SECTION_H_CCY_ROW


def _paginate_sections(sections: list[tuple[str, str, float]]) -> list[list[tuple[str, str, float]]]:
    """
    First-fit por orden: agrupa secciones en hojas respetando PAGE_USABLE_MM.
    Input: lista de (name, html, height_mm).
    Output: lista de páginas, cada página es una lista de secciones.
    """
    pages: list[list[tuple[str, str, float]]] = [[]]
    used = 0.0
    for item in sections:
        _, _, h = item
        if used + h > PAGE_USABLE_MM and pages[-1]:
            pages.append([])
            used = 0.0
        pages[-1].append(item)
        used += h
    return pages


# ── Renderers por sección ────────────────────────────────────────────────────

def _render_kpi_section(agg: dict) -> str:
    d_rv = _delta(agg['revenue'], agg['prev_revenue'])
    d_vl = _delta(agg['volume'], agg['prev_volume'])
    kpi_rev_sub = f"{_fP(d_rv)} vs prev" if agg['prev_revenue'] else "ALL TIME"
    kpi_rev_cls = 'up' if d_rv >= 0 else 'dn'
    kpi_vol_sub = f"{_fP(d_vl)} vs prev" if agg['prev_volume'] else "ALL TIME"
    kpi_vol_cls = 'up' if d_vl >= 0 else 'dn'
    return (
        '<div class="pkr">'
        f'<div class="pkc t"><div class="pkl">REVENUE TODAY</div>'
        f'<div class="pkv t">{_fD(agg["revenue"])}</div>'
        f'<div class="pkd {kpi_rev_cls}">{kpi_rev_sub}</div></div>'
        f'<div class="pkc b"><div class="pkl">VOLUME TODAY</div>'
        f'<div class="pkv b">{_fM(agg["volume"])}</div>'
        f'<div class="pkd {kpi_vol_cls}">{kpi_vol_sub}</div></div>'
        f'<div class="pkc n"><div class="pkl">TRANSFERS</div>'
        f'<div class="pkv">{_f0(agg["transfers"])}</div>'
        f'<div class="pkd">PROFIT: {_fD(agg["profit"])}</div></div>'
        '</div>'
    )


def _render_perf_section(agg: dict) -> str:
    nf_color = '#059669' if agg['net_flow'] >= 0 else '#DC2626'
    nf_label = 'net inflow' if agg['net_flow'] >= 0 else 'net outflow'
    return (
        '<div class="psc">PERFORMANCE SUMMARY</div>'
        '<div class="ppg">'
        f'<div class="ppi"><div class="ppl">AVG REV / TXN</div>'
        f'<div class="ppv">{_fD(agg["avg_rev_per_txn"])}</div>'
        f'<div class="pps">ALL-TIME: {_fD(agg["allrows_avg_rev"])}</div></div>'
        f'<div class="ppi"><div class="ppl">REV / VOL RATIO</div>'
        f'<div class="ppv">{agg["rev_vol_ratio"]*100:.3f}%</div>'
        f'<div class="pps">ALL-TIME: {agg["allrows_rev_vol"]*100:.3f}%</div></div>'
        f'<div class="ppi"><div class="ppl">LARGEST TXN</div>'
        f'<div class="ppv">{_fM(agg["largest_txn"])}</div></div>'
        f'<div class="ppi"><div class="ppl">NET FLOW</div>'
        f'<div class="ppv" style="color:{nf_color}">{_fM(agg["net_flow"])}</div>'
        f'<div class="pps">{nf_label}</div></div>'
        '</div>'
    )


def _render_period_section(agg: dict) -> str:
    d_rv = _delta(agg['revenue'], agg['prev_revenue'])
    d_vl = _delta(agg['volume'], agg['prev_volume'])
    d_tx = _delta(agg['transfers'], agg['prev_transfers'])
    d_rv_color = '#059669' if d_rv >= 0 else '#DC2626'
    d_vl_color = '#059669' if d_vl >= 0 else '#DC2626'
    d_tx_color = '#059669' if d_tx >= 0 else '#DC2626'
    return (
        '<div class="psc">PERIOD COMPARISON</div>'
        '<table class="ptb">'
        '<thead><tr><th style="text-align:left">METRIC</th><th>TOTAL</th>'
        '<th>PREV</th><th>Δ%</th><th>7D AVG</th><th>30D AVG</th><th>ALL-TIME AVG</th></tr></thead>'
        '<tbody>'
        f'<tr><td>REVENUE</td><td style="color:#007A6E;font-weight:700">{_fD(agg["revenue"])}</td>'
        f'<td>{_fD(agg["prev_revenue"])}</td>'
        f'<td style="color:{d_rv_color}">{_fP(d_rv)}</td>'
        f'<td>{_fD(agg["avg_rev_7d"])}</td><td>{_fD(agg["avg_rev_30d"])}</td>'
        f'<td>{_fD(agg["avg_rev_all"])}</td></tr>'
        f'<tr><td>VOLUME</td><td style="color:#1D4ED8;font-weight:700">{_fM(agg["volume"])}</td>'
        f'<td>{_fM(agg["prev_volume"])}</td>'
        f'<td style="color:{d_vl_color}">{_fP(d_vl)}</td>'
        f'<td>{_fM(agg["avg_vol_7d"])}</td><td>{_fM(agg["avg_vol_30d"])}</td>'
        f'<td>{_fM(agg["avg_vol_all"])}</td></tr>'
        f'<tr><td>TRANSFERS</td><td style="font-weight:700">{_f0(agg["transfers"])}</td>'
        f'<td>{agg["prev_transfers"]}</td>'
        f'<td style="color:{d_tx_color}">{_fP(d_tx)}</td>'
        f'<td>{agg["avg_tx_7d"]:.1f}</td><td>{agg["avg_tx_30d"]:.1f}</td>'
        f'<td>{agg["avg_tx_all"]:.1f}</td></tr>'
        '</tbody></table>'
    )


def _render_bank_rails_section(agg: dict) -> str:
    # Banks
    banks_sorted = sorted(agg['by_partner'].items(), key=lambda kv: -kv[1]['rv'])
    max_bank_rv = banks_sorted[0][1]['rv'] if banks_sorted else 1

    def _bank_row(name: str, data: dict) -> str:
        pct = (data['rv'] / agg['revenue'] * 100) if agg['revenue'] else 0
        bar_pct = (data['rv'] / max_bank_rv * 100) if max_bank_rv else 0
        color = BANK_COLORS.get(name, DEFAULT_BANK_COLOR)
        return (
            f'<div class="pbi">'
            f'<div class="pbn">{name}</div>'
            f'<div class="pbr2"><div class="pbf" style="width:{bar_pct:.0f}%;background:{color}"></div></div>'
            f'<div class="pbv">{_fD(data["rv"])}</div>'
            f'<div class="pbp">{pct:.1f}%</div>'
            f'</div>'
        )

    banks_html = ''.join(_bank_row(n, d) for n, d in banks_sorted) or (
        '<div style="color:#7A9CC4;font-size:7.5px">Sin datos en el período</div>'
    )

    # Rails
    types_sorted = sorted(agg['by_type'].items(), key=lambda kv: -kv[1]['rv'])
    rails_html = ''.join(
        f'<div class="prr"><span class="prrl">{t}</span>'
        f'<span class="prrv">{_fD(v["rv"])}</span></div>'
        for t, v in types_sorted
    )

    # Segments (top 5)
    segs_sorted = sorted(agg['by_seg'].items(), key=lambda kv: -kv[1]['rv'])
    top5 = segs_sorted[:5]
    segs_html = ''.join(
        f'<div class="prr"><span class="prrl">{s}</span>'
        f'<span class="prrv">{_fD(v["rv"])}</span></div>'
        for s, v in top5
    )
    top_seg_label = ''
    if top5 and agg['revenue']:
        top = top5[0]
        top_pct = top[1]['rv'] / agg['revenue'] * 100
        top_seg_label = (
            f'<div class="prx">{top[0]}: <strong>{top_pct:.1f}%</strong> of revenue</div>'
        )

    return (
        '<div class="p2c">'
        '<div class="pcd">'
        '<div class="psc" style="margin-top:0">BANK BREAKDOWN</div>'
        f'{banks_html}'
        '</div>'
        '<div class="pcd">'
        '<div class="psc" style="margin-top:0">RAIL &amp; SOCIETY</div>'
        f'{rails_html}'
        '<div class="prsep"></div>'
        f'{segs_html}'
        f'{top_seg_label}'
        '</div>'
        '</div>'
    )


def _render_currency_section(agg: dict) -> str:
    if len(agg['by_ccy']) <= 1:
        return ''
    rows = ''.join(
        f'<tr><td>{c}</td><td>{v["tx"]}</td>'
        f'<td style="color:#007A6E;font-weight:700">{_fD(v["rv"])}</td>'
        f'<td>{_fM(v["vl"])}</td>'
        f'<td>{(v["rv"]/agg["revenue"]*100) if agg["revenue"] else 0:.1f}%</td></tr>'
        for c, v in sorted(agg['by_ccy'].items(), key=lambda kv: -kv[1]['rv'])
    )
    return (
        '<div class="psc">CURRENCY BREAKDOWN</div>'
        '<table class="ptb"><thead><tr>'
        '<th style="text-align:left">CCY</th><th>TXNS</th><th>REVENUE</th>'
        '<th>VOLUME</th><th>% REV</th>'
        '</tr></thead><tbody>' + rows + '</tbody></table>'
    )


def _render_header_block(report_label: str, header_date: str, last_date: str,
                         page_num: int, total_pages: int) -> str:
    page_badge = ''
    if total_pages > 1:
        page_badge = f'<div class="prconf">PÁGINA {page_num} DE {total_pages}</div>'
    return (
        '<div class="ph">'
        '<div class="plogo">'
        '<div class="plogobox">'
        '<svg width="13" height="13" viewBox="0 0 20 20" fill="none">'
        '<circle cx="8" cy="10" r="6.5" fill="#3C3489" opacity=".9"/>'
        '<circle cx="14" cy="10" r="5" fill="#7F77DD" opacity=".65"/>'
        '</svg></div>'
        '<span class="pbr">Ardua</span>'
        '</div>'
        '<div class="pmeta">'
        f'<div class="prtype">{report_label}</div>'
        f'<div class="prdate">{header_date}</div>'
        f'<div class="prconf">DATA AS OF: {last_date}</div>'
        f'{page_badge}'
        '</div>'
        '</div>'
    )


def _render_footer_block(gen_time: str) -> str:
    return (
        '<div class="pft">'
        f'<span>ARDUA OPERATIONS · GENERATED {gen_time}</span>'
        '<span>CONFIDENTIAL · INTERNAL USE ONLY · DO NOT DISTRIBUTE</span>'
        '</div>'
    )


def render_html(agg: dict, report_type: str, generated_at: Optional[datetime] = None) -> str:
    """
    Genera el HTML print-ready del reporte, con paginación automática.

    El Skill decide cuántas hojas A4 necesita según el tamaño estimado del contenido.
    Mismo contenido, mismos estilos que Report 9; solo se reparte en múltiples hojas
    cuando no entra en una sola.
    """
    nw = generated_at or datetime.now()
    report_label = REPORT_LABELS.get(report_type, 'P&L REPORT')
    header_date = nw.strftime('%A, %d %B %Y').upper()
    gen_time = nw.strftime('%d/%m/%Y, %H:%M:%S')

    # Render section fragments
    kpi_html = _render_kpi_section(agg)
    perf_html = _render_perf_section(agg)
    period_html = _render_period_section(agg)
    br_html = _render_bank_rails_section(agg)
    ccy_html = _render_currency_section(agg)

    # Build ordered section list with estimated heights
    sections: list[tuple[str, str, float]] = [
        ('kpis', kpi_html, SECTION_H_KPIS),
        ('perf', perf_html, SECTION_H_PERF),
        ('period', period_html, SECTION_H_PERIOD),
        ('bank_rails', br_html, _estimate_br_h(agg)),
    ]
    if ccy_html:
        sections.append(('currency', ccy_html, _estimate_ccy_h(agg)))

    # Distribute into pages
    pages = _paginate_sections(sections)
    total_pages = len(pages)

    # Render each page as its own .PC block
    pages_body = []
    for i, page_sections in enumerate(pages):
        page_num = i + 1
        header_block = _render_header_block(report_label, header_date,
                                            agg['last_date'], page_num, total_pages)
        body = header_block + ''.join(html for _, html, _ in page_sections)
        body += _render_footer_block(gen_time)
        pages_body.append(f'<div class="PC">{body}</div>')

    return _TEMPLATE.format(pages_body=''.join(pages_body))


# ── Output / entry point ──────────────────────────────────────────────────────

def output_filename(report_type: str, from_d: str, to_d: str) -> str:
    if report_type == 'custom':
        return f"Ardua_PnL_CUSTOM_{from_d}_{to_d}.html"
    return f"Ardua_PnL_{report_type.upper()}_{to_d}.html"


def generate_report(master_path: str | Path,
                    report_type: str = 'daily',
                    from_date: Optional[str] = None,
                    to_date: Optional[str] = None,
                    output_dir: str | Path = '.') -> dict:
    """
    Genera el reporte. Devuelve un dict con:
    - path: path del archivo generado
    - summary: resumen de métricas para mostrar al usuario
    """
    rows = parse_master(master_path)
    if not rows:
        raise PnLError(
            "Después de parsear MASTER.xlsx no quedaron filas válidas. "
            "Revisá que al menos una hoja tenga fechas parseables."
        )

    from_d, to_d = resolve_period(rows, report_type, from_date, to_date)
    agg = compute_aggregates(rows, from_d, to_d)

    html = render_html(agg, report_type)

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = output_filename(report_type, from_d, to_d)
    out_path = out_dir / filename
    out_path.write_text(html, encoding='utf-8')

    # Top segment summary
    segs_sorted = sorted(agg['by_seg'].items(), key=lambda kv: -kv[1]['rv'])
    top_seg = segs_sorted[0] if segs_sorted else (None, {'rv': 0})
    top_pct = (top_seg[1]['rv'] / agg['revenue'] * 100) if (top_seg[0] and agg['revenue']) else 0

    return {
        'path': str(out_path),
        'report_type': report_type,
        'from_date': from_d,
        'to_date': to_d,
        'rows_processed': len(rows),
        'transfers': agg['transfers'],
        'revenue': agg['revenue'],
        'volume': agg['volume'],
        'top_segment': top_seg[0] or '—',
        'top_segment_pct': top_pct,
    }


# ── Template HTML ─────────────────────────────────────────────────────────────

_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Ardua — P&amp;L Report</title>
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;600;700&family=Syne:wght@600;700;800&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
html{{background:#CBD5E1}}
body{{background:#CBD5E1;color:#0D1525;font-family:"DM Sans",sans-serif;-webkit-font-smoothing:antialiased;min-height:100vh;padding:24px 0;display:flex;flex-direction:column;align-items:center;gap:12px}}
@page{{size:A4 portrait;margin:0}}
@media print{{
  html,body{{width:210mm;background:#F2F5FA!important;padding:0!important;display:block!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  .PC{{box-shadow:none!important;margin:0!important;page-break-after:always;break-after:page}}
  .PC:last-child{{page-break-after:auto;break-after:auto}}
}}
.PC{{width:210mm;min-height:297mm;padding:10mm 12mm 8mm;background:#F2F5FA;font-family:"DM Sans",sans-serif;color:#0D1525;font-size:9.5px;box-shadow:0 4px 24px rgba(15,25,45,.18),0 1px 3px rgba(15,25,45,.1);display:flex;flex-direction:column}}
.ph{{display:flex;align-items:center;justify-content:space-between;padding-bottom:5px;border-bottom:2px solid #1A2438;margin-bottom:7px}}
.plogo{{display:flex;align-items:center;gap:6px}}
.plogobox{{width:22px;height:22px;border-radius:5px;background:#0D1525;display:flex;align-items:center;justify-content:center}}
.pbr{{font-family:"Syne",sans-serif;font-size:14px;font-weight:800;color:#0D1525;letter-spacing:-.3px}}
.pmeta{{text-align:right}}
.prtype{{font-size:9px;font-weight:700;color:#0D1525;font-family:"JetBrains Mono",monospace;text-transform:uppercase;letter-spacing:.5px}}
.prdate{{font-size:7px;color:#4A6080;font-family:"JetBrains Mono",monospace;margin-top:1px}}
.prconf{{font-size:6.5px;color:#9CA3AF;font-family:"JetBrains Mono",monospace;margin-top:1px;letter-spacing:.3px}}

.pkr{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:7px}}
.pkc{{background:#fff;border:2px solid #C8D8EE;border-radius:6px;padding:8px 10px}}
.pkc.t{{border-color:#007A6E}}
.pkc.b{{border-color:#1D4ED8}}
.pkc.n{{border-color:#334155}}
.pkl{{font-size:7px;color:#4A6080;text-transform:uppercase;letter-spacing:.6px;font-family:"JetBrains Mono",monospace;font-weight:600;margin-bottom:2px}}
.pkv{{font-family:"JetBrains Mono",monospace;font-size:20px;font-weight:700;line-height:1;color:#0D1525}}
.pkv.t{{color:#007A6E}}
.pkv.b{{color:#1D4ED8}}
.pkd{{font-size:7.5px;font-family:"JetBrains Mono",monospace;margin-top:2px;color:#4A6080}}
.pkd.up{{color:#059669}}
.pkd.dn{{color:#DC2626}}

.psc{{font-size:7.5px;font-weight:700;color:#1A2438;text-transform:uppercase;letter-spacing:.7px;font-family:"JetBrains Mono",monospace;padding-bottom:2px;border-bottom:1px solid #C8D8EE;margin-bottom:4px;margin-top:6px}}

.ppg{{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:5px;margin-bottom:7px}}
.ppi{{background:#fff;border:1px solid #C8D8EE;border-radius:5px;padding:6px 8px}}
.ppl{{font-size:6.5px;color:#4A6080;text-transform:uppercase;font-family:"JetBrains Mono",monospace;letter-spacing:.4px;margin-bottom:1px;font-weight:600}}
.ppv{{font-family:"JetBrains Mono",monospace;font-size:14px;font-weight:700;color:#0D1525;line-height:1}}
.pps{{font-size:7px;color:#7A9CC4;font-family:"JetBrains Mono",monospace;margin-top:1px}}

.ptb{{width:100%;border-collapse:collapse;font-size:8px;margin-bottom:5px}}
.ptb th{{background:#1A2438;color:#fff;font-family:"JetBrains Mono",monospace;font-size:6px;text-transform:uppercase;letter-spacing:.4px;padding:3px 5px;text-align:center;font-weight:700}}
.ptb td{{padding:3px 5px;border-bottom:1px solid #E4ECF8;font-family:"JetBrains Mono",monospace;text-align:right;background:#fff;color:#334155}}
.ptb td:first-child{{text-align:left;font-weight:600;color:#0D1525;background:#EDF2FA}}
.ptb tr:nth-child(even) td{{background:#F5F8FF}}
.ptb tr:nth-child(even) td:first-child{{background:#E8EFF8}}

.p2c{{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:6px}}
.pcd{{background:#fff;border:1px solid #C8D8EE;border-radius:5px;padding:6px 8px}}

.pbi{{display:flex;align-items:center;gap:4px;margin-bottom:2px;font-size:7.5px;font-family:"JetBrains Mono",monospace}}
.pbn{{width:55px;font-weight:700;color:#0D1525;flex-shrink:0}}
.pbr2{{flex:1;height:4px;background:#E4ECF8;border-radius:2px;overflow:hidden}}
.pbf{{height:100%;border-radius:2px}}
.pbv{{width:60px;text-align:right;font-weight:700;color:#0D1525}}
.pbp{{width:34px;text-align:right;color:#7A9CC4;font-size:7px}}

.prr{{display:flex;justify-content:space-between;padding:2px 0;font-size:7.5px;font-family:"JetBrains Mono",monospace}}
.prrl{{font-weight:700;color:#0D1525}}
.prrv{{color:#007A6E;font-weight:700}}
.prsep{{border-top:1px solid #C8D8EE;margin:3px 0}}
.prx{{font-size:6.5px;color:#7A9CC4;font-family:"JetBrains Mono",monospace;margin-top:2px}}

.pft{{border-top:1px solid #C8D8EE;padding-top:3px;margin-top:auto;display:flex;justify-content:space-between;font-size:6px;color:#7A9CC4;font-family:"JetBrains Mono",monospace}}
</style>
</head>
<body>
{pages_body}
</body>
</html>
'''
